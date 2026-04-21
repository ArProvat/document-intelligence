from pathlib import Path

from openpyxl import load_workbook

from app.models.schemas import ExtractionMethod, PageResult
from app.processors.text_cleaner import clean_text


def _stringify_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def process_xlsx(xlsx_path: Path) -> list[PageResult]:
    workbook = load_workbook(filename=str(xlsx_path), data_only=True, read_only=True)
    pages: list[PageResult] = []

    for page_number, sheet in enumerate(workbook.worksheets, start=1):
        rows: list[list[str]] = []

        for row in sheet.iter_rows(values_only=True):
            cells = [_stringify_cell(value) for value in row]
            if any(cells):
                rows.append(cells)

        text_lines = [f"Sheet: {sheet.title}"]
        text_lines.extend(" | ".join(cell for cell in row if cell) for row in rows)
        raw_text = "\n".join(line for line in text_lines if line.strip())
        cleaned_text, warnings = clean_text(raw_text)

        if not rows:
            warnings.append(f"Sheet '{sheet.title}' is empty")

        pages.append(
            PageResult(
                page_number=page_number,
                raw_text=raw_text,
                cleaned_text=cleaned_text,
                confidence=1.0 if cleaned_text.strip() else 0.0,
                extraction_method=ExtractionMethod.NATIVE_TEXT,
                word_count=len(cleaned_text.split()),
                has_tables=bool(rows),
                tables=[rows] if rows else [],
                warnings=warnings,
            )
        )

    workbook.close()
    return pages
